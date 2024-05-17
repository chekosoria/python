"""Aplicación para comparar dos reportes"""
import tkinter as tk
import logging
from tkinter import filedialog, messagebox, END
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageTk
import xlrd


def configurar_logger():
    """Función para configurar log"""
    SISTEMA = "Genos"

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    # Formato del log
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    # Manejador para el archivo principal
    file_handler = logging.FileHandler(f"{SISTEMA.lower()}.log")
    file_handler.setFormatter(formatter)

    # Manejador para la consola
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)

    # Añadir los manejadores al logger
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)

    return logger


class CompareReport(tk.Frame):
    """Pantalla para comparar reportes"""

    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.logger = logging.getLogger(__name__)
        self.logger.info("Inicializando pantalla para comparar reportes")

        # Crear contenedor principal para centrar los elementos
        container = tk.Frame(self)
        container.pack(expand=True, fill="both")

        # Crear título
        label_titulo = tk.Label(
            container, text="Comparar reportes", font=("Helvetica", 24))
        label_titulo.grid(row=0, column=0, columnspan=2, pady=20)

        # Ruta del directorio actual
        current_directory = os.path.dirname(os.path.realpath(__file__))

        # Cargar y convertir el icono
        icon_path = os.path.join(current_directory, "search.png")
        search_image = Image.open(icon_path)
        search_icon = ImageTk.PhotoImage(search_image)
        icon_exe_path = os.path.join(current_directory, "play.png")
        exe_image = Image.open(icon_exe_path)
        exe_icon = ImageTk.PhotoImage(exe_image)
        icon_clean_path = os.path.join(current_directory, "clean.png")
        clean_image = Image.open(icon_clean_path)
        clean_icon = ImageTk.PhotoImage(clean_image)

        # Widgets para buscar los reportes a comparar
        self.label_base = tk.Label(
            container, text="Seleccione reporte base", font=("Helvetica", 12, "bold"))
        self.entry_base = tk.Entry(container, width=30)
        self.button_search_base = tk.Button(
            container, image=search_icon, command=lambda: self.load_file(self.entry_base))
        self.button_search_base.image = search_icon
        self.label_compare = tk.Label(
            container, text="Seleccione reporte a comparar", font=("Helvetica", 12, "bold"))
        self.entry_compare = tk.Entry(container, width=30)
        self.button_search_compare = tk.Button(
            container, image=search_icon, command=lambda: self.load_file(self.entry_compare))
        self.button_search_compare.image = search_icon
        self.button_exe = tk.Button(
            container, image=exe_icon, command=lambda: self.compare_files())
        self.button_exe.image = exe_icon
        self.button_clean = tk.Button(
            container, image=clean_icon, command=lambda: self.clean())
        self.button_clean.image = clean_icon

        # Ubicar los widgets en el contenedor
        self.label_base.grid(row=1, columnspan=2, padx=5,
                             pady=5, sticky="nsew")
        self.entry_base.grid(row=2, column=0,
                             padx=5, pady=5, sticky="e")
        self.button_search_base.grid(row=2, column=1, pady=5, sticky="w")
        self.label_compare.grid(row=3, columnspan=2,
                                padx=5, pady=5, sticky="nsew")
        self.entry_compare.grid(row=4, column=0,
                                padx=5, pady=5, sticky="e")
        self.button_search_compare.grid(row=4, column=1, pady=5, sticky="w")
        self.button_exe.grid(row=5, column=0, pady=5, sticky="e")
        self.button_clean.grid(
            row=5, column=1, pady=5, sticky="w")

        # Centrar el contenedor en la ventana principal
        container.grid_rowconfigure(0, weight=0)
        container.grid_rowconfigure(1, weight=0)
        container.grid_rowconfigure(2, weight=0)
        container.grid_rowconfigure(3, weight=0)
        container.grid_rowconfigure(4, weight=0)
        container.grid_rowconfigure(5, weight=0)
        container.grid_columnconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=1)

    def load_file(self, entry):
        """Método para cargar archivo"""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")])
        if filepath:
            entry.delete(0, tk.END)
            entry.insert(0, filepath)

    def convert(self, file_path):
        """Método para convertir archivos XLS a XLSX"""
        if file_path.endswith('.xls'):
            # Leer archivo .xls
            xls_data = pd.read_excel(file_path, engine='xlrd')
            # Crear un nuevo archivo .xlsx
            new_file_path = file_path.replace('.xls', '.xlsx')
            xls_data.to_excel(new_file_path, index=False, engine='openpyxl')
            return new_file_path
        return file_path

    def clean(self):
        """Método para limpiar formulario"""

        # Limpiar las entradas de texto
        self.entry_base.delete(0, END)
        self.entry_compare.delete(0, END)

    def compare_files(self):
        """Método para comparar reportes"""
        file1_path = self.entry_base.get()
        file2_path = self.entry_compare.get()

        if not file1_path or not file2_path:
            messagebox.showerror(
                "Error", "Por favor seleccione ambos reportes.")
            return

        try:
            base_name = os.path.splitext(os.path.basename(file1_path))[0]

            # Leer archivos Excel con xlrd
            book1 = xlrd.open_workbook(file1_path)
            book2 = xlrd.open_workbook(file2_path)

            # Obtener las hojas de cálculo
            sheet1 = book1.sheet_by_index(0)
            sheet2 = book2.sheet_by_index(0)

            # Crear un nuevo libro de trabajo para el resultado
            wb_result = openpyxl.Workbook()
            ws_result = wb_result.active

            # Colores para resaltar diferencias
            highlight = PatternFill(start_color='FFFF00',
                                    end_color='FFFF00', fill_type='solid')

            differences_found = False  # Bandera para indicar si se encontraron diferencias

            # Recorrer todas las celdas y comparar
            for row in range(max(sheet1.nrows, sheet2.nrows)):
                for col in range(max(sheet1.ncols, sheet2.ncols)):
                    value1 = sheet1.cell_value(row, col)
                    value2 = sheet2.cell_value(row, col)

                    if value1 != value2:
                        differences_found = True
                        result_cell = ws_result.cell(row=row+1, column=col+1)
                        result_cell.value = f"{value1} -> {value2}"
                        result_cell.fill = highlight
                    else:
                        result_cell = ws_result.cell(row=row+1, column=col+1)
                        result_cell.value = value1

            if differences_found:
                # Guardar el resultado
                result_path_with_format = f"Comparativo_{base_name}.xlsx"
                wb_result.save(result_path_with_format)
                messagebox.showinfo("Éxito",
                                    f"Revisión completada!\nLos resultados se han guardado en:\n{
                                        result_path_with_format}")
                self.logger.info("Se han encontrado diferencias, los resultados se han guardado en %s",
                                 result_path_with_format)
            else:
                messagebox.showinfo("Sin diferencias",
                                    "No se encontraron diferencias entre los reportes.")
                self.logger.info(
                    "No se encontraron diferencias entre los reportes.")

        except (pd.errors.ParserError, openpyxl.utils.exceptions.InvalidFileException) as e:
            messagebox.showerror(
                "Error", "Error al leer archivos Excel: " + str(e))
            self.logger.info("Error al leer archivos Excel.")
