"""Aplicación para probar endpoint y comparar archivos a la vez"""
import os
import subprocess
import zipfile
import shutil
from datetime import datetime, timedelta
import logging
import tkinter as tk
from tkinter import ttk, messagebox, END
import csv
from PIL import Image, ImageTk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import xlrd
from utils.db_utils import inicializar_base_de_datos, obtener_conexion


def configurar_logger():
    """Función para configurar log"""
    sistema = "Genos"

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    # Formato del log
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    # Manejador para el archivo principal
    file_handler = logging.FileHandler(f"{sistema.lower()}.log")
    file_handler.setFormatter(formatter)

    # Manejador para la consola
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)

    # Añadir los manejadores al logger
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)

    return logger


class BatchTest(tk.Frame):
    """Pantalla para probar Endpoints"""

    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        self.logger = logging.getLogger(__name__)
        self.logger.info(
            "Inicializando pantalla para probar Endpoints y comparar reportes")

        # Crear contenedor principal para centrar los elementos
        container = tk.Frame(self)
        container.pack(expand=True, fill="both")

        # Crear título
        label_titulo = tk.Label(
            container, text="Prueba completa de Endpoints", font=("Helvetica", 24, "bold"))
        label_titulo.grid(row=0, column=0, columnspan=2, pady=20)

        # Ruta del directorio actual
        current_directory = os.path.dirname(os.path.realpath(__file__))

        # Cargar y convertir el icono
        icon_exe_path = os.path.join(current_directory, "play.png")
        exe_image = Image.open(icon_exe_path)
        exe_icon = ImageTk.PhotoImage(exe_image)
        icon_clean_path = os.path.join(current_directory, "clean.png")
        clean_image = Image.open(icon_clean_path)
        clean_icon = ImageTk.PhotoImage(clean_image)

        # Crear botón con icono
        self.button_exe = tk.Button(
            container, image=exe_icon, command=self.test_endpoints)
        self.button_clean = tk.Button(
            container, image=clean_icon, command=self.clean_screen)
        # Referencia para evitar que se recolecte basura
        self.button_exe.image = exe_icon
        self.button_clean.image = clean_icon

        # Widget para mostrar el estado de la prueba
        self.label_status = tk.Label(
            container, text="Estado:", font=("Helvetica", 12, "bold"))
        self.status_text = tk.Text(container, height=25, width=70)
        self.status_text.configure(bg="black", fg="white")

        # Widgets para ingresar alias, URL y parámetros del endpoint
        self.label_first_endpoint = tk.Label(
            container, text="Seleccione el Endpoint base:", font=("Helvetica", 10, "bold"))
        self.combo_first_endpoint = ttk.Combobox(container,
                                                 width=30,
                                                 )
        self.label_second_endpoint = tk.Label(
            container, text="Seleccione el Endpint a comparar:", font=("Helvetica", 10, "bold"))
        self.combo_second_endpoint = ttk.Combobox(container,
                                                  width=30,
                                                  )

        # Ubicar los widgets en el contenedor
        self.label_first_endpoint.grid(row=3, column=0, sticky="e")
        self.combo_first_endpoint.grid(row=3, column=1, sticky="w")
        self.label_second_endpoint.grid(row=4, column=0, sticky="e")
        self.combo_second_endpoint.grid(row=4, column=1, sticky="w")
        self.button_exe.grid(row=5, column=0, columnspan=2, pady=5)
        self.label_status.grid(row=6, column=0, sticky="w")
        self.status_text.grid(row=7, column=0, columnspan=2,
                              padx=5, pady=5, sticky="nsew")
        self.button_clean.grid(
            row=8, column=0, columnspan=2, pady=5)

        # Centrar el contenedor en la ventana principal
        container.grid_rowconfigure(0, weight=0)
        container.grid_rowconfigure(1, weight=0)
        container.grid_rowconfigure(2, weight=0)
        container.grid_rowconfigure(3, weight=0)
        container.grid_rowconfigure(4, weight=0)
        container.grid_rowconfigure(5, weight=0)
        container.grid_rowconfigure(6, weight=0)
        container.grid_rowconfigure(7, weight=0)
        container.grid_rowconfigure(8, weight=0)
        container.grid_columnconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=1)

        # Inicializar base de datos
        inicializar_base_de_datos()

        # Vincular el método cargar_endpoints al evento de mostrar la pantalla
        self.bind("<<ShowFrame>>", self.on_show_frame)

    def on_show_frame(self, event):
        """Cargar los endpoints existentes desde la base de datos al mostrar la pantalla"""
        self.cargar_endpoints()

    def cargar_endpoints(self):
        """Cargar los endpoints existentes desde la base de datos y actualiza los comboboxes"""
        aliases = []
        with obtener_conexion() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT alias FROM endpoints")
            rows = cursor.fetchall()
            aliases = [row[0] for row in rows]

        # Actualizar los valores de los comboboxes
        self.combo_first_endpoint['values'] = aliases
        self.combo_second_endpoint['values'] = aliases

        self.combo_first_endpoint.set('')  # Limpiar selección actual
        self.combo_second_endpoint.set('')  # Limpiar selección actual

    def test_endpoints(self):
        """Función para probar endpoints"""
        base_alias = self.combo_first_endpoint.get()
        compare_alias = self.combo_second_endpoint.get()

        if not base_alias or not compare_alias:
            messagebox.showerror(
                "Error", "Seleccione ambos endpoints para probar.")
            return

        # Comparar los archivos generados
        base_dir = self.run_endpoint_test(base_alias, "_base")
        compare_dir = self.run_endpoint_test(compare_alias, "")
        result_dir = "test_results"

        # Comparar los archivos generados
        self.compare_files(base_dir, compare_dir, result_dir)

    def run_endpoint_test(self, alias, suffix):
        """Ejecutar prueba para un endpoint específico y descargar el archivo generado"""
        with obtener_conexion() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT url, parametros, download_url FROM endpoints WHERE alias = ?", (alias,))
            endpoint_data = cursor.fetchone()
            if not endpoint_data:
                self.status_text.insert(tk.END, f"No se encontró el endpoint {
                                        alias} en la base de datos.\n")
                return None

            url, params, download = endpoint_data
            start_time = datetime.now()
            self.status_text.insert(
                END, f"Iniciando prueba del endpoint {alias}...\n")
            self.status_text.see(tk.END)
            self.master.update()
            retry_count = 3  # Número máximo de reintentos
            retry = 0
            success = False
            generated_filename = None

            while retry < retry_count and not success:
                try:
                    created_file = f"{alias}{suffix}.txt"
                    command = ["curl", "-o", created_file, "-k", url + params]
                    process = subprocess.Popen(
                        command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
                    stdout, stderr = process.communicate()
                    if process.returncode == 0:
                        end_time = datetime.now()
                        elapsed_time_seconds = (
                            end_time - start_time).total_seconds()
                        elapsed_timedelta = timedelta(
                            seconds=elapsed_time_seconds)
                        elapsed_time_formatted = str(elapsed_timedelta)

                        with open(created_file, "r", encoding="latin-1") as f:
                            generated_filename = f.readline().strip()

                        self.status_text.insert(
                            END, f"{datetime.now(
                            )} - Se ha generado el reporte {
                                alias} en {elapsed_time_formatted} segundos\n"
                        )
                        self.save_result_to_csv(
                            alias, start_time, end_time,
                            elapsed_time_formatted, "Completado", generated_filename)
                        success = True

                        if success:
                            download_command = [
                                "curl", "-O", "-k", download + generated_filename]
                            process = subprocess.Popen(
                                download_command,
                                stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
                            stdout, stderr = process.communicate()
                            if process.returncode == 0:
                                end_time = datetime.now()
                                elapsed_time_seconds = (
                                    end_time - start_time).total_seconds()
                                elapsed_timedelta = timedelta(
                                    seconds=elapsed_time_seconds)
                                elapsed_time_formatted = str(elapsed_timedelta)

                                self.status_text.insert(
                                    END, f"{datetime.now()} - Se ha descargado el reporte {
                                        generated_filename} en {elapsed_time_formatted} segundos\n"
                                )
                                self.save_result_to_csv(
                                    alias, start_time, end_time,
                                    elapsed_time_formatted, "Descargado", generated_filename)

                                with zipfile.ZipFile(generated_filename, 'r') as zip_ref:
                                    extract_dir = f"endpoint{suffix}"
                                    if not os.path.exists(extract_dir):
                                        os.makedirs(extract_dir)
                                    for member in zip_ref.namelist():
                                        filename = os.path.basename(member)
                                        if filename:
                                            source = zip_ref.open(member)
                                            target = open(os.path.join(
                                                extract_dir, filename), "wb")
                                            with source, target:
                                                shutil.copyfileobj(
                                                    source, target)
                                    self.status_text.insert(END,
                                                            f"Reportes extraidos en {
                                                                extract_dir}...\n")
                                    self.logger.info(
                                        "Reportes extraidos en %s", extract_dir)

                            else:
                                self.status_text.insert(END, f"Error al descargar el reporte {
                                                        generated_filename}: {stderr.decode()}\n")
                            self.logger.info(
                                "Reporte %s descargado.", generated_filename)

                        if os.path.exists(created_file):
                            os.remove(created_file)
                    else:
                        self.status_text.insert(END, f"Error al generar el reporte {
                                                alias}: {stderr.decode()}\n")
                        retry += 1
                        if retry < retry_count:
                            self.status_text.insert(
                                END, f"Reintentando ({retry}/{retry_count})...\n")
                except Exception as e:
                    self.status_text.insert(
                        END, f"Error al generar el reporte {alias}: {str(e)}\n")
                    retry += 1
                    if retry < retry_count:
                        self.status_text.insert(
                            END, f"Reintentando ({retry}/{retry_count})...\n")
            if not success:
                self.status_text.insert(tk.END, f"No se pudo generar el reporte {
                                        alias} después de {retry_count} intentos.\n")
                return None

            return extract_dir

    def save_result_to_csv(self, alias, start_time, end_time,
                           elapsed_time_formatted, status, generated_filename):
        """Función para almacenar el resultado de la prueba en un CSV"""
        current_date = datetime.now().strftime("%Y-%m-%d")
        csv_filename = f"endpoint_log_{current_date}.csv"
        with open(csv_filename, 'a', newline='', encoding="latin-1") as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow([alias, start_time.strftime("%H:%M:%S"), end_time.strftime(
                "%H:%M:%S"), elapsed_time_formatted, status, generated_filename])
        self.logger.info("Registro de prueba guardado en %s: %s, %s, %s, %s, %s",
                         csv_filename, alias, start_time, end_time, elapsed_time_formatted, status)

    def compare_files(self, base_dir, compare_dir, result_dir):
        """Método para comparar archivos descomprimidos de dos carpetas y 
        generar archivos de comparativos en caso de diferencias"""
        if not os.path.exists(base_dir) or not os.path.exists(compare_dir):
            messagebox.showerror(
                "Error", "No se generaron ambos reportes para comparar.")
            return

        if not os.path.exists(result_dir):
            os.makedirs(result_dir)

        for base_file in os.listdir(base_dir):
            base_file_path = os.path.join(base_dir, base_file)
            compare_file_path = os.path.join(compare_dir, base_file)

            if os.path.exists(compare_file_path) and base_file.endswith('.xls'):
                try:
                    self.status_text.insert(
                        END, f"Iniciando comparativo de {
                            base_file_path} contra {compare_file_path}\n")

                    base_name = os.path.splitext(base_file)[0]

                    # Leer archivos Excel
                    book1 = xlrd.open_workbook(base_file_path)
                    book2 = xlrd.open_workbook(compare_file_path)

                    sheet1 = book1.sheet_by_index(0)
                    sheet2 = book2.sheet_by_index(0)

                    wb_result = openpyxl.Workbook()
                    ws_result = wb_result.active

                    highlight = PatternFill(
                        start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    differences_found = False

                    for row in range(max(sheet1.nrows, sheet2.nrows)):
                        for col in range(max(sheet1.ncols, sheet2.ncols)):
                            value1 = sheet1.cell_value(row, col)
                            value2 = sheet2.cell_value(row, col)

                            if value1 != value2:
                                differences_found = True
                                result_cell = ws_result.cell(
                                    row=row+1, column=col+1)
                                result_cell.value = f"{value1} -> {value2}"
                                result_cell.fill = highlight
                            else:
                                result_cell = ws_result.cell(
                                    row=row+1, column=col+1)
                                result_cell.value = value1

                    if differences_found:
                        result_path_with_format = os.path.join(
                            result_dir, f"Comparativo_{base_name}.xlsx")
                        wb_result.save(result_path_with_format)
                        self.status_text.insert(END, f"Se han encontrado diferencias en {
                                                base_file_path}. Resultados guardados en {
                                                    result_path_with_format}\n")
                        self.logger.info(
                            "Se han encontrado diferencias en %s. Resultados guardados en %s",
                            base_file_path, result_path_with_format)
                    else:
                        self.status_text.insert(
                            END, f"No se encontraron diferencias entre {base_file_path} y {compare_file_path}\n")
                        self.logger.info(
                            "No se encontraron diferencias entre %s y %s", base_file_path, compare_file_path)

                except (pd.errors.ParserError, openpyxl.utils.exceptions.InvalidFileException) as e:
                    messagebox.showerror(
                        "Error", "Error al leer archivos Excel: " + str(e))
                    self.logger.info("Error al leer archivos Excel.")

        messagebox.showinfo("Éxito", "Prueba completada.")
        self.logger.info("Prueba completada.")

    def clean_screen(self):
        """Limpiar los Combobox y el Text widget status_text"""
        self.combo_first_endpoint.set('')  # Limpiar Combobox
        self.combo_second_endpoint.set('')  # Limpiar Combobox
        self.status_text.delete('1.0', tk.END)  # Limpiar Text widget
